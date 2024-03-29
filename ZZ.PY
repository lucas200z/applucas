import openpyxl
import pandas as pd

# Caminho para o arquivo Excel
caminho_excel = 'PRODUTOS2024.xlsx'

# Número de linhas a serem lidas em cada chunk
tamanho_chunk = 10

# Função para processar cada chunk
def processar_chunk(chunk):
    # Seu código de processamento aqui
    print(chunk)

# Abrir o arquivo Excel usando openpyxl
workbook = openpyxl.load_workbook(caminho_excel)
sheet = workbook.active

# Número total de linhas no arquivo
total_linhas = sheet.max_row

# Variável para armazenar o início do próximo chunk
inicio_chunk = 1

# Iterar sobre os chunks
while inicio_chunk <= total_linhas:
    # Calcular o fim do chunk
    fim_chunk = min(inicio_chunk + tamanho_chunk - 1, total_linhas)

    # Ler o chunk usando openpyxl
    chunk_data = []
    for row in sheet.iter_rows(min_row=inicio_chunk, max_row=fim_chunk, values_only=True):
        chunk_data.append(row)

    # Criar um DataFrame pandas a partir do chunk
    chunk_df = pd.DataFrame(chunk_data, columns=sheet[1])

    # Processar o chunk
    processar_chunk(chunk_df)

    # Atualizar o início para o próximo chunk
    inicio_chunk = fim_chunk + 1
